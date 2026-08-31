"""Client-facing reports: the daily digest and the weekly REPORT.xlsx."""

from __future__ import annotations

import argparse
import json
import shutil
import subprocess
from datetime import date as Date, timedelta
from pathlib import Path
from typing import Any, Iterable, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from contracts import CONTRACTS, Field
from export import display, read_records
from publish import SOURCES

# Kata yang tidak boleh muncul di teks yang ditulis pipeline untuk klien
# (docs/CLIENT_REPORT.md §2 aturan 3 dan fase P11 langkah 2).
FORBIDDEN_WORDS = (
    "selector",
    "selectolax",
    "tenacity",
    "xpath",
    "stacktrace",
    "traceback",
    "regex",
    "checkpoint",
    "sqlite",
    "exception",
    "storage_state",
    "semaphore",
)
MONTHS = (
    "Januari", "Februari", "Maret", "April", "Mei", "Juni",
    "Juli", "Agustus", "September", "Oktober", "November", "Desember",
)
MAX_EXAMPLES = 3  # docs/CLIENT_REPORT.md §2 aturan 2
WEEK_DAYS = 7
HEALTHY = "SEHAT"
ATTENTION = "PERLU PERHATIAN"
# Tiga warna saja: hijau sehat, kuning warning, merah critical. Tidak ada warna lain.
FILLS = {
    "sehat": PatternFill("solid", fgColor="C6EFCE"),
    "warning": PatternFill("solid", fgColor="FFEB9C"),
    "critical": PatternFill("solid", fgColor="FFC7CE"),
}
MAX_COLUMN_WIDTH = 45


class JargonError(ValueError):
    """Raised when text meant for the client carries pipeline vocabulary."""


# ----------------------------------------------------------------- validator


def jargon_found(text: str, quoted: Iterable[Any] = ()) -> list[str]:
    """Forbidden words in prose the pipeline wrote, ignoring quoted source data.

    Nilai yang dikutip apa adanya dari situs sumber bukan jargon kami: halaman
    HTTPX benar-benar berjudul "Exceptions". Yang dijaga adalah kalimat yang
    ditulis pipeline sendiri.
    """
    lowered = text.casefold()
    for value in quoted:
        rendered = str(value).casefold().strip()
        if rendered:
            lowered = lowered.replace(rendered, " ")
    return sorted({word for word in FORBIDDEN_WORDS if word in lowered})


def assert_no_jargon(text: str, what: str, quoted: Iterable[Any] = ()) -> None:
    found = jargon_found(text, quoted)
    if found:
        raise JargonError(f"{what} memuat jargon terlarang: {', '.join(found)}")


# ------------------------------------------------------------------- helpers


def source_name(target: str) -> str:
    if target in SOURCES:
        return SOURCES[target]["name"]
    return "driftlab (fixture uji lokal)"


def number(value: Any, decimals: int = 0) -> str:
    if value is None:
        return "—"
    text = f"{float(value):,.{decimals}f}"
    return text.replace(",", "\x00").replace(".", ",").replace("\x00", ".")


def percent(value: Any) -> str:
    return "—" if value is None else f"{number(float(value) * 100, 1)}%"


def long_date(iso: str) -> str:
    day = Date.fromisoformat(iso)
    return f"{day.day} {MONTHS[day.month - 1]} {day.year}"


def duration_text(seconds: Any) -> str:
    if seconds is None:
        return "—"
    total = int(round(float(seconds)))
    minutes, rest = divmod(total, 60)
    return f"{minutes} menit {rest} detik" if minutes else f"{rest} detik"


def clock(timestamp: Any) -> str:
    return str(timestamp)[11:16] if timestamp else "—"


def load_json(path: Path) -> dict[str, Any] | None:
    return json.loads(path.read_text(encoding="utf-8")) if path.is_file() else None


def load_alerts(reports_root: Path) -> list[dict[str, Any]]:
    path = reports_root / "alerts.jsonl"
    if not path.is_file():
        return []
    with path.open(encoding="utf-8") as source:
        return [json.loads(line) for line in source if line.strip()]


def alerts_for(alerts: Sequence[dict[str, Any]], target: str, run_date: str) -> list[dict[str, Any]]:
    """Alarm yang masih berlaku. Baris ber-`resolved_at` sudah ditutup run berikutnya —
    mengutipnya membuat laporan klien membantah tabel angkanya sendiri."""
    return [
        a
        for a in alerts
        if a.get("target") == target and a.get("date") == run_date and not a.get("resolved_at")
    ]


def status_of(alerts: Sequence[dict[str, Any]]) -> str:
    return ATTENTION if alerts else HEALTHY


def severity_of(alerts: Sequence[dict[str, Any]]) -> str:
    if any(alert.get("severity") == "critical" for alert in alerts):
        return "critical"
    return "warning" if alerts else "sehat"


def contract_fields(target: str) -> list[Field]:
    return list(CONTRACTS[target]["fields"])


def field_label(target: str, name: str) -> str:
    for field in contract_fields(target):
        if field.name == name:
            return field.description
    return name


def required_completeness(run: dict[str, Any] | None) -> float | None:
    if not run:
        return None
    completeness = run.get("field_completeness", {})
    required = [field.name for field in contract_fields(run.get("target", "")) if field.required]
    values = [completeness[name] for name in required if name in completeness]
    return min(values) if values else None


def record_label(record: dict[str, Any] | None) -> str:
    if not record:
        return ""
    fields = record.get("fields", {})
    for name in ("title", "name", "author_name", "url"):
        if fields.get(name) not in (None, "", []):
            return str(fields[name])
    return str(record.get("record_id", ""))


def snapshot_records(data_root: Path, target: str, run_date: str) -> dict[str, dict[str, Any]]:
    path = data_root / target / run_date / "records.jsonl"
    if not path.is_file():
        return {}
    return {record["record_id"]: record for record in read_records(path)}


def available_dates(data_root: Path, target: str) -> list[str]:
    directory = data_root / target
    if not directory.is_dir():
        return []
    return sorted(item.name for item in directory.iterdir() if (item / "run.json").is_file())


# --------------------------------------------------------------- daily digest


def daily_context(target: str, run_date: str, data_root: Path, reports_root: Path) -> dict[str, Any]:
    diff = load_json(reports_root / target / run_date / "diff.json") or {
        "counts": {"added": 0, "changed": 0, "removed": 0, "baseline_total": 0},
        "added": [], "changed": [], "removed": [], "baseline_date": None,
    }
    baseline_date = diff.get("baseline_date")
    run_today = load_json(data_root / target / run_date / "run.json")
    run_baseline = load_json(data_root / target / baseline_date / "run.json") if baseline_date else None
    diff_baseline = (
        load_json(reports_root / target / baseline_date / "diff.json") if baseline_date else None
    ) or {}
    return {
        "target": target,
        "date": run_date,
        "baseline_date": baseline_date,
        "diff": diff,
        "diff_baseline": diff_baseline,
        "run_today": run_today,
        "run_baseline": run_baseline,
        "records": snapshot_records(data_root, target, run_date),
        "alerts": alerts_for(load_alerts(reports_root), target, run_date),
    }


def render_daily(context: dict[str, Any]) -> tuple[str, list[str]]:
    """Return the digest markdown plus the verbatim source values it quotes."""
    target = context["target"]
    diff = context["diff"]
    counts = diff.get("counts", {})
    previous = context["diff_baseline"].get("counts", {})
    run_today = context["run_today"]
    alerts = context["alerts"]
    quoted: list[str] = []

    status = status_of(alerts)
    mark = "✅" if status == HEALTHY else "⚠️"
    lines = [
        f"# {source_name(target)} — {long_date(context['date'])}",
        "",
        f"**Status: {status}** {mark}",
        "",
        "| | Hari ini | Kemarin |",
        "|---|---|---|",
        f"| Total data | {number((run_today or {}).get('records_unique'))} "
        f"| {number((context['run_baseline'] or {}).get('records_unique'))} |",
        f"| Baru | {number(counts.get('added'))} | {number(previous.get('added'))} |",
        f"| Berubah | {number(counts.get('changed'))} | {number(previous.get('changed'))} |",
        f"| Hilang | {number(counts.get('removed'))} | {number(previous.get('removed'))} |",
        f"| Kelengkapan kolom | {percent(required_completeness(run_today))} "
        f"| {percent(required_completeness(context['run_baseline']))} |",
        "",
    ]

    if alerts:
        lines.append(f"**Yang perlu Anda tahu ({len(alerts)})**:")
        for alert in alerts:
            lines.append(f"- {alert.get('message', '')}")
        lines.append("")

    added = diff.get("added", [])
    lines.append(f"**Yang baru ({number(counts.get('added'))})**" + _examples_suffix(added))
    for entry in added[:MAX_EXAMPLES]:
        summary = str(entry.get("summary", ""))
        quoted.append(summary)
        lines.append(f"- \"{summary}\" — [lihat]({entry.get('url')})")
    if not added:
        lines.append("- tidak ada data baru hari ini")
    lines.append("")

    changed = diff.get("changed", [])
    lines.append(f"**Yang berubah ({number(counts.get('changed'))})**" + _examples_suffix(changed))
    for entry in changed[:MAX_EXAMPLES]:
        label = record_label(context["records"].get(entry.get("record_id")))
        quoted.append(label)
        for change in entry.get("fields_changed", []):
            before = display(change.get("from"))
            after = display(change.get("to"))
            quoted.extend([before, after])
            lines.append(
                f"- \"{label}\": {field_label(target, change.get('field', ''))} "
                f"{before or '(kosong)'} → {after or '(kosong)'}"
            )
    if not changed:
        lines.append("- tidak ada perubahan nilai hari ini")
    lines.append("")

    removed = diff.get("removed", [])
    lines.append(f"**Yang hilang ({number(counts.get('removed'))})**" + _examples_suffix(removed))
    for entry in removed[:MAX_EXAMPLES]:
        summary = str(entry.get("summary", ""))
        quoted.append(summary)
        last_seen = entry.get("last_seen")
        seen = long_date(last_seen) if last_seen else "sebelumnya"
        lines.append(f"- \"{summary}\" — terakhir terlihat {seen}")
    if not removed:
        lines.append("- tidak ada data yang hilang hari ini")
    lines.append("")

    lines.append(_pipeline_note(run_today))
    lines.append(_rate_limit_note(run_today))
    lines.append("")
    return "\n".join(lines), quoted


def _examples_suffix(entries: Sequence[Any]) -> str:
    if len(entries) > MAX_EXAMPLES:
        return f" — {MAX_EXAMPLES} contoh teratas:"
    return ":"


def _pipeline_note(run_today: dict[str, Any] | None) -> str:
    if not run_today:
        return (
            "**Catatan pipeline:** pengambilan data terjadwal hari ini tidak menghasilkan run. "
            "Data hari sebelumnya tetap utuh dan aman."
        )
    errors = int(run_today.get("errors", 0))
    tail = "tanpa error" if errors == 0 else f"{number(errors)} permintaan gagal"
    return (
        f"**Catatan pipeline:** run selesai {clock(run_today.get('finished_at'))}, "
        f"{duration_text(run_today.get('duration_sec'))}, {tail}."
    )


def _rate_limit_note(run_today: dict[str, Any] | None) -> str:
    limit = (run_today or {}).get("rate_limit", {})
    agreed = limit.get("delay_sec")
    agreed_text = f"{number(agreed, 2)} detik" if agreed is not None else "1,00 detik"
    gap = limit.get("observed_min_gap_ms")
    if gap is None:
        return (
            f"Tidak ada permintaan baru ke situs sumber hari ini; kesepakatan jeda "
            f"{agreed_text} per permintaan tetap berlaku."
        )
    return (
        f"Jeda terpendek antar permintaan {number(float(gap) / 1000, 2)} detik "
        f"(sesuai kesepakatan {agreed_text} per permintaan)."
    )


def build_daily(target: str, run_date: str, data_root: Path, reports_root: Path) -> Path:
    context = daily_context(target, run_date, data_root, reports_root)
    markdown, quoted = render_daily(context)
    assert_no_jargon(markdown, f"daily.md {target} {run_date}", quoted)
    output = reports_root / target / run_date / "daily.md"
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(markdown, encoding="utf-8")
    return output


# --------------------------------------------------------------- notification


def notification_text(target: str, alert: dict[str, Any], baseline_date: str | None) -> str:
    """Four client lines, always in the same order (docs/CLIENT_REPORT.md §4)."""
    severity = alert.get("severity", "critical")
    weight = (
        "kritis — data hari ini belum layak dipakai sampai saya periksa"
        if severity == "critical"
        else "peringatan — data masih bisa dipakai, tetapi saya awasi"
    )
    safe = "Data hari sebelumnya tetap utuh dan aman."
    if baseline_date:
        safe = f"Data {long_date(baseline_date)} tetap utuh dan aman."
    body = [
        str(alert.get("message", "")),
        f"Tingkat keparahan: {weight}.",
        str(alert.get("likely_cause", "")),
        f"Saya sudah mulai memeriksa; perkiraan perbaikan: hari ini. {safe}",
    ]
    header = f"[DriftWatch] {source_name(target)} — {ATTENTION}"
    text = header + "\n\n" + "\n".join(body) + "\n"
    assert_no_jargon(text, f"notifikasi {target} {alert.get('code')}")
    return text


def send_notifications(
    target: str, run_date: str, data_root: Path, reports_root: Path, notify: bool = True
) -> list[Path]:
    diff = load_json(reports_root / target / run_date / "diff.json") or {}
    baseline_date = diff.get("baseline_date")
    written: list[Path] = []
    directory = reports_root / target / run_date
    directory.mkdir(parents=True, exist_ok=True)
    for alert in alerts_for(load_alerts(reports_root), target, run_date):
        if alert.get("severity") != "critical":
            continue
        text = notification_text(target, alert, baseline_date)
        output = directory / f"notification-{alert['code']}.txt"
        output.write_text(text, encoding="utf-8")
        written.append(output)
        if notify and shutil.which("notify-send"):
            subprocess.run(["notify-send", text.splitlines()[0], "\n".join(text.splitlines()[2:])], check=False)
    return written


# --------------------------------------------------------------- weekly xlsx


def union_fields(targets: Sequence[str]) -> list[Field]:
    seen: dict[str, Field] = {}
    for target in targets:
        for field in contract_fields(target):
            seen.setdefault(field.name, field)
    return list(seen.values())


def week_dates(end_date: str) -> list[str]:
    last = Date.fromisoformat(end_date)
    return [(last - timedelta(days=offset)).isoformat() for offset in range(WEEK_DAYS - 1, -1, -1)]


def _finish(sheet, widths: Sequence[int]) -> None:
    sheet.freeze_panes = "A2"
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    for index, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _write(sheet, header: Sequence[str], rows: Sequence[Sequence[Any]]) -> None:
    sheet.append(list(header))
    for row in rows:
        sheet.append(list(row))
    widths = []
    for column in range(len(header)):
        longest = max([len(str(header[column]))] + [len(str(row[column])) for row in rows] or [0])
        widths.append(min(max(longest + 2, 10), MAX_COLUMN_WIDTH))
    _finish(sheet, widths)


def build_weekly(
    end_date: str,
    targets: Sequence[str],
    data_root: Path,
    reports_root: Path,
    output: Path,
) -> tuple[Path, dict[str, int]]:
    dates = week_dates(end_date)
    alerts = load_alerts(reports_root)
    book = Workbook()

    summary = book.active
    summary.title = "Ringkasan"
    summary_rows: list[list[Any]] = []
    health_rows: list[list[Any]] = []
    severities: list[str] = []
    change_rows: list[list[Any]] = []
    new_rows: list[list[Any]] = []
    fields = union_fields(targets)

    for run_date in dates:
        scheduled = [target for target in targets if (data_root / target).is_dir()]
        succeeded = [
            target for target in scheduled
            if (load_json(data_root / target / run_date / "run.json") or {}).get("exit_code") == 0
        ]
        uptime = len(succeeded) / len(scheduled) if scheduled else None
        for target in targets:
            run = load_json(data_root / target / run_date / "run.json")
            diff = load_json(reports_root / target / run_date / "diff.json")
            if run is None and diff is None:
                continue
            day_alerts = alerts_for(alerts, target, run_date)
            severities.append(severity_of(day_alerts))
            counts = (diff or {}).get("counts", {})
            summary_rows.append([
                run_date, target,
                (run or {}).get("records_unique"),
                counts.get("added"), counts.get("changed"), counts.get("removed"),
                required_completeness(run),
                uptime,
                len(day_alerts),
                status_of(day_alerts),
            ])
            statuses = (run or {}).get("http_status_counts", {})
            requests = sum(int(value) for value in statuses.values())
            errors = sum(
                int(value) for status, value in statuses.items() if not str(status).startswith("2")
            )
            completeness = (run or {}).get("field_completeness", {})
            health_rows.append([
                run_date, target,
                clock((run or {}).get("finished_at")),
                (run or {}).get("duration_sec"),
                requests, errors,
                ", ".join(sorted(alert["code"] for alert in day_alerts)) or "—",
                status_of(day_alerts),
            ] + [completeness.get(field.name) for field in fields])

            records = snapshot_records(data_root, target, run_date)
            for entry in (diff or {}).get("changed", []):
                label = record_label(records.get(entry.get("record_id")))
                for change in entry.get("fields_changed", []):
                    change_rows.append([
                        run_date, target, label,
                        field_label(target, change.get("field", "")),
                        display(change.get("from")), display(change.get("to")),
                        entry.get("url"),
                    ])
            for entry in (diff or {}).get("added", []):
                record = records.get(entry.get("record_id"), {})
                values = record.get("fields", {})
                new_rows.append(
                    [run_date, target, entry.get("summary"), entry.get("url")]
                    + [display(values.get(field.name)) if field.name in values else "" for field in fields]
                )

    _write(
        summary,
        ["Tanggal", "Target", "Total data", "Baru", "Berubah", "Hilang",
         "Kelengkapan kolom", "Uptime pipeline", "Jumlah alarm", "Status"],
        summary_rows,
    )
    _paint_status(summary, 10, severities)
    _percent_format(summary, columns=(7, 8))

    _write(
        book.create_sheet("Perubahan"),
        ["Tanggal", "Target", "Nama", "Kolom", "Dari", "Ke", "Tautan"],
        change_rows,
    )
    _write(
        book.create_sheet("Data Baru"),
        ["Tanggal", "Target", "Nama", "Tautan"] + [field.name for field in fields],
        new_rows,
    )
    health = book.create_sheet("Kesehatan Pipeline")
    _write(
        health,
        ["Tanggal", "Target", "Jam run", "Durasi (detik)", "Jumlah permintaan", "Error",
         "Alarm", "Status"] + [f"Kelengkapan {field.name}" for field in fields],
        health_rows,
    )
    _paint_status(health, 8, severities)
    _percent_format(health, columns=tuple(range(9, 9 + len(fields))))

    dictionary_rows = [
        [target, field.name, _type_name(field.type),
         _example(data_root, target, dates, field.name),
         field.source_hint, "wajib" if field.required else "opsional", field.description]
        for target in targets
        for field in contract_fields(target)
    ]
    _write(
        book.create_sheet("Kamus Data"),
        ["Target", "Kolom", "Tipe", "Contoh", "Sumber di halaman", "Wajib/Opsional", "Catatan"],
        dictionary_rows,
    )

    output.parent.mkdir(parents=True, exist_ok=True)
    book.save(output)
    return output, {
        "Ringkasan": len(summary_rows),
        "Perubahan": len(change_rows),
        "Data Baru": len(new_rows),
        "Kesehatan Pipeline": len(health_rows),
        "Kamus Data": len(dictionary_rows),
    }


def _paint_status(sheet, column: int, severities: Sequence[str]) -> None:
    """Three fills only: green healthy, yellow warning, red critical."""
    for offset, severity in enumerate(severities):
        sheet.cell(row=2 + offset, column=column).fill = FILLS[severity]


def _percent_format(sheet, columns: Sequence[int]) -> None:
    for column in columns:
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row=row, column=column).number_format = "0.0%"


def _type_name(kind: Any) -> str:
    names = {str: "teks", int: "bilangan bulat", float: "angka desimal", bool: "ya/tidak", list: "daftar"}
    return names.get(kind, getattr(kind, "__name__", str(kind)))


def _example(data_root: Path, target: str, dates: Sequence[str], name: str) -> str:
    for run_date in reversed(dates):
        path = data_root / target / run_date / "records.jsonl"
        if not path.is_file():
            continue
        for record in read_records(path):
            value = record.get("fields", {}).get(name)
            if value not in (None, "", []):
                return display(value)[:80]
    return ""


# ---------------------------------------------------------------------- CLI


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--target", choices=list(CONTRACTS))
    group = parser.add_mutually_exclusive_group()
    group.add_argument("--today", action="store_true")
    group.add_argument("--date")
    parser.add_argument("--all", action="store_true", help="Semua tanggal yang punya data")
    parser.add_argument("--weekly", action="store_true", help="Bangun REPORT.xlsx tujuh hari")
    parser.add_argument("--notify", action="store_true", help="Kirim notifikasi alarm critical")
    parser.add_argument("--no-send", action="store_true", help="Tulis notifikasi tanpa mengirim")
    parser.add_argument("--targets", nargs="+", choices=list(CONTRACTS), default=list(CONTRACTS))
    parser.add_argument("--data-root", type=Path, default=Path("data"))
    parser.add_argument("--reports-root", type=Path, default=Path("reports"))
    parser.add_argument("--out", type=Path, help="Tujuan REPORT.xlsx")
    args = parser.parse_args()

    run_date = Date.today().isoformat() if args.today else args.date

    if args.weekly:
        output = args.out or (args.reports_root / "REPORT.xlsx")
        path, rows = build_weekly(
            run_date or Date.today().isoformat(), args.targets, args.data_root, args.reports_root, output
        )
        print(f"{path}: " + " · ".join(f"{name} {count} baris" for name, count in rows.items()))
        return 0

    targets = [args.target] if args.target else list(args.targets)
    written: list[Path] = []
    for target in targets:
        dates = available_dates(args.data_root, target) if args.all else [run_date]
        for day in dates:
            if day is None:
                parser.error("pilih --today, --date, atau --all")
            try:
                written.append(build_daily(target, day, args.data_root, args.reports_root))
                if args.notify:
                    written.extend(
                        send_notifications(target, day, args.data_root, args.reports_root, not args.no_send)
                    )
            except JargonError as error:
                print(f"BUILD GAGAL {target} {day}: {error}")
                return 2
    for path in written:
        print(path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
