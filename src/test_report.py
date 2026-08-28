"""Unit tests for the client digest, its jargon gate, and the weekly workbook."""

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

import report
from contracts import CONTRACTS


def manifest(target="books", records=1000, duration=12.0):
    return {
        "target": target,
        "run_id": "2026-08-28T09-00-00",
        "records_unique": records,
        "exit_code": 0,
        "duration_sec": duration,
        "finished_at": "2026-08-28T09:18:41+07:00",
        "errors": 0,
        "field_completeness": {field.name: 1.0 for field in CONTRACTS[target]["fields"]},
        "http_status_counts": {"200": records},
        "schema_unknown_fields": [],
        "rate_limit": {"delay_sec": 1.0, "observed_min_gap_ms": 1000},
    }


def record(record_id, title, url="https://books.toscrape.com/x"):
    return {
        "record_id": record_id,
        "target": "books",
        "url": url,
        "content_hash": "sha256:x",
        "fields": {field.name: 1 for field in CONTRACTS["books"]["fields"]} | {"title": title, "upc": record_id},
        "missing_fields": [],
        "missing_reason": {},
    }


class Workspace:
    """Minimal data/ + reports/ tree that report.py can be pointed at."""

    def __init__(self, root: Path):
        self.data = root / "data"
        self.reports = root / "reports"

    def snapshot(self, target, run_date, records, run=None):
        directory = self.data / target / run_date
        directory.mkdir(parents=True, exist_ok=True)
        (directory / "run.json").write_text(json.dumps(run or manifest(target)), encoding="utf-8")
        with (directory / "records.jsonl").open("w", encoding="utf-8") as output:
            for item in records:
                output.write(json.dumps(item, ensure_ascii=False) + "\n")

    def diff(self, target, run_date, payload):
        directory = self.reports / target / run_date
        directory.mkdir(parents=True, exist_ok=True)
        (directory / "diff.json").write_text(json.dumps(payload), encoding="utf-8")

    def alert(self, payload):
        self.reports.mkdir(parents=True, exist_ok=True)
        with (self.reports / "alerts.jsonl").open("a", encoding="utf-8") as output:
            output.write(json.dumps(payload, ensure_ascii=False) + "\n")


def empty_diff(baseline_date="2026-08-27", **counts):
    return {
        "target": "books", "date": "2026-08-28", "baseline_date": baseline_date,
        "counts": {"added": 0, "changed": 0, "removed": 0, "unchanged": 0, "baseline_total": 1000, **counts},
        "added": [], "changed": [], "removed": [],
        "health": {"error_rate": 0.0}, "alarms": [],
    }


class JargonGateTests(unittest.TestCase):
    def test_forbidden_word_in_pipeline_prose_fails_the_build(self):
        with tempfile.TemporaryDirectory() as temporary:
            space = Workspace(Path(temporary))
            space.snapshot("books", "2026-08-28", [record("a1", "Buku Satu")])
            space.diff("books", "2026-08-28", empty_diff(baseline_date=None))
            space.alert({
                "target": "books", "date": "2026-08-28", "code": "RUN_FAILED", "severity": "critical",
                "message": "Pengambilan berhenti dengan traceback panjang.",
                "likely_cause": "x", "next_action": "y",
            })

            with self.assertRaises(report.JargonError) as raised:
                report.build_daily("books", "2026-08-28", space.data, space.reports)

            self.assertIn("traceback", str(raised.exception))
            self.assertFalse((space.reports / "books/2026-08-28/daily.md").exists())

    def test_forbidden_word_quoted_from_the_source_site_is_not_jargon(self):
        # Halaman HTTPX benar-benar berjudul "Exceptions"; itu data klien, bukan jargon kami.
        self.assertEqual(report.jargon_found('- "Exceptions - HTTPX"', ["Exceptions - HTTPX"]), [])
        self.assertEqual(report.jargon_found('- "Exceptions - HTTPX"'), ["exception"])

    def test_every_forbidden_word_is_detected(self):
        for word in report.FORBIDDEN_WORDS:
            self.assertEqual(report.jargon_found(f"kalimat {word} di sini"), [word])


class DailyDigestTests(unittest.TestCase):
    def build(self, space, **kwargs):
        return report.build_daily("books", "2026-08-28", space.data, space.reports, **kwargs)

    def test_silent_day_still_produces_a_digest_with_status_and_rate_limit(self):
        with tempfile.TemporaryDirectory() as temporary:
            space = Workspace(Path(temporary))
            space.snapshot("books", "2026-08-27", [record("a1", "Buku Satu")])
            space.snapshot("books", "2026-08-28", [record("a1", "Buku Satu")])
            space.diff("books", "2026-08-28", empty_diff(unchanged=1000))

            path = self.build(space)
            lines = path.read_text(encoding="utf-8").splitlines()

            self.assertEqual(lines[0], "# books.toscrape.com — 28 Agustus 2026")
            self.assertEqual(lines[2], "**Status: SEHAT** ✅")
            self.assertIn("| Baru | 0 | — |", lines)
            self.assertTrue(any("kesepakatan" in line and "permintaan" in line for line in lines))

    def test_at_most_three_examples_per_category_and_from_to_values(self):
        with tempfile.TemporaryDirectory() as temporary:
            space = Workspace(Path(temporary))
            records = [record(f"a{index}", f"Buku {index}") for index in range(6)]
            space.snapshot("books", "2026-08-28", records)
            payload = empty_diff(added=6, changed=1, removed=4)
            payload["added"] = [
                {"record_id": item["record_id"], "url": item["url"], "summary": item["fields"]["title"]}
                for item in records
            ]
            payload["changed"] = [{
                "record_id": "a0", "url": records[0]["url"],
                "fields_changed": [{"field": "price_incl_tax_gbp", "from": 51.77, "to": 48.5}],
            }]
            payload["removed"] = [
                {"record_id": f"z{index}", "url": "u", "summary": f"Lama {index}", "last_seen": "2026-08-27"}
                for index in range(4)
            ]
            space.diff("books", "2026-08-28", payload)

            text = self.build(space).read_text(encoding="utf-8")

            self.assertEqual(text.count("— [lihat]("), report.MAX_EXAMPLES)
            self.assertEqual(text.count("terakhir terlihat"), report.MAX_EXAMPLES)
            self.assertIn("Harga termasuk pajak dalam GBP 51.77 → 48.5", text)
            self.assertIn("**Yang baru (6)** — 3 contoh teratas:", text)


class NotificationTests(unittest.TestCase):
    def test_template_is_four_lines_in_the_locked_order(self):
        alert = {
            "code": "RECORD_COUNT_DROP", "severity": "critical",
            "message": "Hanya 412 data terkumpul, kurang dari 80% jumlah normal 988.",
            "likely_cause": "Sebagian halaman daftar kemungkinan tidak lagi terbaca.",
            "next_action": "Jalankan `make recon` untuk target ini.",
        }

        text = report.notification_text("books", alert, "2026-08-27")
        header, blank, *body = text.splitlines()

        self.assertEqual(header, "[DriftWatch] books.toscrape.com — PERLU PERHATIAN")
        self.assertEqual(blank, "")
        self.assertEqual(len(body), 4)
        self.assertEqual(body[0], alert["message"])
        self.assertTrue(body[1].startswith("Tingkat keparahan: kritis"))
        self.assertEqual(body[2], alert["likely_cause"])
        self.assertIn("Data 27 Agustus 2026 tetap utuh dan aman.", body[3])
        # next_action milik developer (`make recon`) tidak pernah dikirim ke klien.
        self.assertNotIn("make recon", text)

    def test_only_critical_alerts_are_sent(self):
        with tempfile.TemporaryDirectory() as temporary:
            space = Workspace(Path(temporary))
            space.snapshot("books", "2026-08-28", [record("a1", "Buku Satu")])
            space.diff("books", "2026-08-28", empty_diff())
            space.alert({
                "target": "books", "date": "2026-08-28", "code": "DURATION_ANOMALY", "severity": "warning",
                "message": "Run hari ini lebih lama dari biasanya.", "likely_cause": "c", "next_action": "n",
            })
            space.alert({
                "target": "books", "date": "2026-08-28", "code": "ZERO_RECORDS", "severity": "critical",
                "message": "Tidak ada data yang berhasil dikumpulkan hari ini.",
                "likely_cause": "Tata letak sumber mungkin berubah.", "next_action": "n",
            })

            written = report.send_notifications("books", "2026-08-28", space.data, space.reports, notify=False)

            self.assertEqual([path.name for path in written], ["notification-ZERO_RECORDS.txt"])


class WeeklyWorkbookTests(unittest.TestCase):
    def test_five_sheets_dictionary_from_contracts_and_three_status_colors(self):
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            space = Workspace(root)
            space.snapshot("books", "2026-08-26", [record("a1", "Buku Satu")])
            space.snapshot("books", "2026-08-27", [record("a1", "Buku Satu")])
            space.snapshot("books", "2026-08-28", [record("a1", "Buku Satu")])
            space.diff("books", "2026-08-26", empty_diff(baseline_date=None))
            space.diff("books", "2026-08-27", empty_diff(baseline_date="2026-08-26"))
            space.diff("books", "2026-08-28", empty_diff())
            space.alert({
                "target": "books", "date": "2026-08-27", "code": "DURATION_ANOMALY", "severity": "warning",
                "message": "m", "likely_cause": "c", "next_action": "n",
            })
            space.alert({
                "target": "books", "date": "2026-08-28", "code": "ZERO_RECORDS", "severity": "critical",
                "message": "m", "likely_cause": "c", "next_action": "n",
            })

            output, rows = report.build_weekly(
                "2026-08-28", ["books"], space.data, space.reports, root / "REPORT.xlsx"
            )
            book = load_workbook(output)

            self.assertEqual(
                book.sheetnames,
                ["Ringkasan", "Perubahan", "Data Baru", "Kesehatan Pipeline", "Kamus Data"],
            )
            self.assertEqual(rows["Kamus Data"], len(CONTRACTS["books"]["fields"]))
            summary = book["Ringkasan"]
            self.assertEqual(summary.freeze_panes, "A2")
            colors = [summary.cell(row, 10).fill.fgColor.rgb for row in range(2, summary.max_row + 1)]
            self.assertEqual(colors, ["00C6EFCE", "00FFEB9C", "00FFC7CE"])
            dictionary = book["Kamus Data"]
            self.assertEqual(
                [dictionary.cell(row, 2).value for row in range(2, dictionary.max_row + 1)],
                [field.name for field in CONTRACTS["books"]["fields"]],
            )
            self.assertEqual(dictionary.cell(2, 4).value, "a1")


if __name__ == "__main__":
    unittest.main()
